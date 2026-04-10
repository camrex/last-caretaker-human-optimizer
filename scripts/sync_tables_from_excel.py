#!/usr/bin/env python3
"""Sync app data tables from the Excel workbook.

Usage:
  python scripts/sync_tables_from_excel.py --write
  python scripts/sync_tables_from_excel.py --check
"""

from __future__ import annotations

import argparse
import json
from datetime import time
from pathlib import Path
from typing import Any

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
WORKBOOK_PATH = ROOT / "resource" / "The Last Caretaker - Human Needs.xlsx"
MAP_WORKBOOK_PATH = ROOT / "resource" / "The Last Caretaker - Map Locations.xlsx"
FOODS_PATH = ROOT / "data" / "foods.js"
MEMORIES_PATH = ROOT / "data" / "memories.js"
PROFESSIONS_PATH = ROOT / "data" / "professions.js"
MAP_LOCATIONS_PATH = ROOT / "data" / "map_locations.js"
MAP_CONNECTIONS_PATH = ROOT / "data" / "map_connections.js"
ITEM_IDS_PATH = ROOT / "resource" / "item_ids.json"

FOOD_NUMERIC_KEYS = [
    "lifeExp",
    "height",
    "weight",
    "strength",
    "intellect",
    "carbs",
    "protein",
    "fat",
    "omega3",
    "vitd",
    "calcium",
    "mito",
    "nanite",
    "bio",
]

MEMORY_NUMERIC_KEYS = [
    "adaptability",
    "comms",
    "creativity",
    "discipline",
    "empathy",
    "focus",
    "leadership",
    "logic",
    "patience",
    "wisdom",
    "starChild",
]

PROF_NUMERIC_KEYS = [
    "lifeExp",
    "height",
    "weight",
    "strength",
    "intellect",
    "comms",
    "empathy",
    "leadership",
    "discipline",
    "focus",
    "adaptability",
    "creativity",
    "patience",
    "wisdom",
    "logic",
    "starChild",
]


def clean_num(v: Any) -> int:
    return 0 if v is None else int(v)


def clean_coord(v: Any) -> float:
    if v is None or v == "":
        return 0.0
    try:
        return float(v)
    except Exception:
        return 0.0


def clean_bool(v: Any) -> bool:
    if isinstance(v, bool):
        return v
    if v is None:
        return False
    text = str(v).strip().lower()
    return text in {"1", "true", "yes", "y", "x"}


def rank_to_num(rank_value: Any) -> int:
    if rank_value is None:
        return 0
    text = str(rank_value)
    try:
        return int(text.split("-")[0])
    except Exception:
        return 0


def time_to_text(v: Any) -> str:
    if isinstance(v, time):
        return f"{v.hour}:{v.minute:02d}"
    if isinstance(v, (int, float)) and 0 <= float(v) < 1:
        total_minutes = int(round(float(v) * 24 * 60))
        hours, minutes = divmod(total_minutes, 60)
        return f"{hours}:{minutes:02d}"
    if v is None:
        return ""
    return str(v)


def js_quote(v: Any) -> str:
    return json.dumps(v, ensure_ascii=True)


def load_item_ids_registry() -> dict[str, dict[str, int]]:
    if not ITEM_IDS_PATH.exists():
        return {"foods": {}, "memories": {}, "professions": {}}

    raw = json.loads(ITEM_IDS_PATH.read_text(encoding="utf-8"))
    out = {"foods": {}, "memories": {}, "professions": {}}
    for key in out:
        src = raw.get(key, {}) if isinstance(raw, dict) else {}
        if not isinstance(src, dict):
            continue
        for name, ident in src.items():
            try:
                n = int(ident)
            except Exception:
                continue
            if n > 0:
                out[key][str(name)] = n
    return out


def assign_stable_ids(items: list[dict[str, Any]], kind: str, registry: dict[str, dict[str, int]]) -> bool:
    mapping = registry.setdefault(kind, {})
    existing_ids = {int(v) for v in mapping.values() if isinstance(v, int) and v > 0}
    next_id = (max(existing_ids) + 1) if existing_ids else 1
    changed = False

    for item in items:
        name = str(item.get("name") or "").strip()
        if not name:
            continue
        if name in mapping:
            item["id"] = int(mapping[name])
            continue
        mapping[name] = next_id
        item["id"] = next_id
        next_id += 1
        changed = True

    return changed


def render_item_ids_registry(registry: dict[str, dict[str, int]]) -> str:
    clean: dict[str, dict[str, int]] = {}
    for key in ["foods", "memories", "professions"]:
        clean[key] = dict(sorted(
            ((name, int(ident)) for name, ident in registry.get(key, {}).items()),
            key=lambda pair: pair[1]
        ))
    return json.dumps(clean, indent=2, ensure_ascii=True) + "\n"


def parse_foods(wb: openpyxl.Workbook) -> list[dict[str, Any]]:
    ws = wb["FOOD"]
    foods: list[dict[str, Any]] = []
    for row in ws.iter_rows(min_row=4, values_only=True):
        if not row or not row[0]:
            continue
        if row[0] == "Food":
            continue
        foods.append(
            {
                "name": str(row[0]).strip(),
                "rank": rank_to_num(row[1]),
                "level": clean_num(row[2]),
                "craftTime": time_to_text(row[3]),
                "lifeExp": clean_num(row[4]),
                "height": clean_num(row[5]),
                "weight": clean_num(row[6]),
                "strength": clean_num(row[7]),
                "intellect": clean_num(row[8]),
                "carbs": clean_num(row[9]),
                "protein": clean_num(row[10]),
                "fat": clean_num(row[11]),
                "omega3": clean_num(row[12]),
                "vitd": clean_num(row[13]),
                "calcium": clean_num(row[14]),
                "mito": clean_num(row[15]),
                "nanite": clean_num(row[16]),
                "bio": clean_num(row[17]),
                "tooltip": str(row[18]).strip() if row[18] else "",
            }
        )
    return foods


def parse_memories(wb: openpyxl.Workbook) -> list[dict[str, Any]]:
    ws = wb["MEMORIES"]
    memories: list[dict[str, Any]] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        memories.append(
            {
                "name": str(row[0]).strip(),
                "rank": rank_to_num(row[1]),
                "adaptability": clean_num(row[2]),
                "comms": clean_num(row[3]),
                "creativity": clean_num(row[4]),
                "discipline": clean_num(row[5]),
                "empathy": clean_num(row[6]),
                "focus": clean_num(row[7]),
                "leadership": clean_num(row[8]),
                "logic": clean_num(row[9]),
                "patience": clean_num(row[10]),
                "wisdom": clean_num(row[11]),
                "starChild": clean_num(row[12]),
                "tooltip": str(row[13]).strip() if len(row) > 13 and row[13] else "",
            }
        )
    return memories


def parse_professions(wb: openpyxl.Workbook) -> list[dict[str, Any]]:
    ws = wb["PROFESSION"]
    professions: list[dict[str, Any]] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[1]:
            continue
        if str(row[1]).startswith("*"):
            continue
        committee = None if row[4] is None else str(row[4]).strip()
        professions.append(
            {
                "name": str(row[1]).strip(),
                "tier": clean_num(row[2]),
                "committee": committee,
                "lifeExp": clean_num(row[6]),
                "height": clean_num(row[7]),
                "weight": clean_num(row[8]),
                "strength": clean_num(row[9]),
                "intellect": clean_num(row[10]),
                "comms": clean_num(row[11]),
                "empathy": clean_num(row[12]),
                "leadership": clean_num(row[13]),
                "discipline": clean_num(row[14]),
                "focus": clean_num(row[15]),
                "adaptability": clean_num(row[16]),
                "creativity": clean_num(row[17]),
                "patience": clean_num(row[18]),
                "wisdom": clean_num(row[19]),
                "logic": clean_num(row[20]),
                "starChild": clean_num(row[21]),
            }
        )
    return professions


def parse_map_locations(path: Path) -> list[dict[str, Any]]:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["LOCATIONS"] if "LOCATIONS" in wb.sheetnames else wb[wb.sheetnames[0]]
    locations: list[dict[str, Any]] = []

    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())
    header_idx = {str(value).strip().upper(): idx for idx, value in enumerate(header_row) if value is not None}

    def cell_value(row: tuple[Any, ...], column_name: str, fallback_index: int | None = None) -> Any:
        idx = header_idx.get(column_name)
        if idx is None:
            idx = fallback_index
        if idx is None or idx >= len(row):
            return None
        return row[idx]

    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=1):
        x_val = cell_value(row, "X", 0)
        y_val = cell_value(row, "Y", 1)
        name_val = cell_value(row, "NAME", 2)
        if not row or x_val is None or y_val is None or not name_val:
            continue

        name = str(name_val).strip()
        type_name = str(cell_value(row, "TYPE", 3) or "Unknown").strip()
        note = str(cell_value(row, "NOTE", 4) or "").strip()
        radius = clean_num(cell_value(row, "RADIUS", 5))
        spoiler_flag = clean_bool(cell_value(row, "SPOILER", 6))
        spoiler_trigger = str(cell_value(row, "SPOILER_TRIGGER", 7) or "").strip()
        locations.append(
            {
                "id": idx,
                "x": clean_coord(x_val),
                "y": clean_coord(y_val),
                "name": name,
                "type": type_name,
                "note": note,
                "radius": radius,
                # Keep current UI spoiler behavior limited to explicit unknown names for now.
                "spoiler": name == "[UNKNOWN]",
                # Future-ready spreadsheet metadata for trigger-driven reveal logic.
                "spoilerFlag": spoiler_flag,
                "spoilerTrigger": spoiler_trigger,
            }
        )

    return locations


def parse_map_connections(path: Path) -> list[dict[str, Any]]:
    wb = openpyxl.load_workbook(path, data_only=True)
    if "CONNECTIONS" not in wb.sheetnames:
        return []

    ws = wb["CONNECTIONS"]
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())
    header_idx = {str(value).strip().upper(): idx for idx, value in enumerate(header_row) if value is not None}

    def cell_value(row: tuple[Any, ...], column_name: str, fallback_index: int | None = None) -> Any:
        idx = header_idx.get(column_name)
        if idx is None:
            idx = fallback_index
        if idx is None or idx >= len(row):
            return None
        return row[idx]

    connections: list[dict[str, Any]] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row:
            continue
        ident = clean_num(cell_value(row, "ID", 0))
        from_name = str(cell_value(row, "FROM", 1) or "").strip()
        to_name = str(cell_value(row, "TO", 4) or "").strip()
        if not from_name or not to_name:
            continue

        connections.append(
            {
                "id": ident,
                "from": from_name,
                "fromX": clean_coord(cell_value(row, "FROM_X", 2)),
                "fromY": clean_coord(cell_value(row, "FROM_Y", 3)),
                "to": to_name,
                "toX": clean_coord(cell_value(row, "TO_X", 5)),
                "toY": clean_coord(cell_value(row, "TO_Y", 6)),
            }
        )

    return connections


def render_sparse_object(item: dict[str, Any], fixed: list[str], numeric: list[str]) -> str:
    parts: list[str] = []
    for key in fixed:
        value = item[key]
        if value is None:
            parts.append(f"{key}:null")
        elif isinstance(value, str):
            parts.append(f"{key}:{js_quote(value)}")
        else:
            parts.append(f"{key}:{value}")

    for key in numeric:
        value = int(item.get(key, 0) or 0)
        if value != 0:
            parts.append(f"{key}:{value}")

    if item.get("tooltip"):
        parts.append(f"tooltip:{js_quote(item['tooltip'])}")

    return "{ " + ", ".join(parts) + " }"


def render_foods(foods: list[dict[str, Any]]) -> str:
    lines = [
        "// data/foods.js",
        "// Food definitions for The Last Caretaker optimizer.",
        "// Sparse format: only non-zero numeric traits are stored.",
        "",
        "var FOODS = [",
    ]
    for i, item in enumerate(foods):
        text = "  " + render_sparse_object(item, ["id", "name", "rank", "craftTime", "level"], FOOD_NUMERIC_KEYS)
        if i < len(foods) - 1:
            text += ","
        lines.append(text)
    lines.extend([
        "];",
        "",
        "// Foods that require Bio Flesh ingredients (mito/nanite/bio)",
        "var BIO_FLESH_FOODS = "
        + js_quote([f["name"] for f in foods if (f.get("mito", 0) + f.get("nanite", 0) + f.get("bio", 0)) > 0])
        + ";",
        "",
    ])
    return "\n".join(lines)


def render_memories(memories: list[dict[str, Any]]) -> str:
    lines = [
        "// data/memories.js",
        "// Memory definitions for The Last Caretaker optimizer.",
        "// Sparse format: only non-zero numeric traits are stored.",
        "",
        "var MEMORIES = [",
    ]
    for i, item in enumerate(memories):
        text = "  " + render_sparse_object(item, ["id", "name", "rank"], MEMORY_NUMERIC_KEYS)
        if i < len(memories) - 1:
            text += ","
        lines.append(text)
    lines.extend([
        "];",
        "",
        "// Artifact memories flagged for the exclude toggle",
        "var ARTIFACT_MEMORIES = " + js_quote([m["name"] for m in memories if m.get("rank") == 5]) + ";",
        "",
    ])
    return "\n".join(lines)


def render_professions(professions: list[dict[str, Any]]) -> str:
    lines = [
        "// data/professions.js",
        "// Profession definitions for The Last Caretaker optimizer.",
        "// Sparse format: only non-zero numeric traits are stored.",
        "",
        "var PROFESSIONS = [",
    ]
    for i, item in enumerate(professions):
        text = "  " + render_sparse_object(item, ["id", "name", "tier", "committee"], PROF_NUMERIC_KEYS)
        if i < len(professions) - 1:
            text += ","
        lines.append(text)
    lines.extend([
        "];",
        "",
    ])
    return "\n".join(lines)


def render_map_locations(locations: list[dict[str, Any]]) -> str:
    def js_number(v: Any) -> str:
        try:
            num = float(v)
        except Exception:
            return "0"
        if num.is_integer():
            return str(int(num))
        return format(num, "g")

    lines = [
        "// data/map_locations.js",
        "// Map locations synced from The Last Caretaker - Map Locations.xlsx.",
        "// Coordinates use 100m grid units; negative Y values are north.",
        "// `spoilerFlag` and `spoilerTrigger` are future-facing metadata from the workbook.",
        "",
        "var MAP_LOCATIONS = [",
    ]
    for i, item in enumerate(locations):
        parts = [
            f"id:{item['id']}",
            f"x:{js_number(item['x'])}",
            f"y:{js_number(item['y'])}",
            f"name:{js_quote(item['name'])}",
            f"type:{js_quote(item['type'])}",
        ]
        if item.get("note"):
            parts.append(f"note:{js_quote(item['note'])}")
        if int(item.get("radius", 0) or 0) > 0:
            parts.append(f"radius:{int(item['radius'])}")
        if item.get("spoiler"):
            parts.append("spoiler:true")
        if item.get("spoilerFlag"):
            parts.append("spoilerFlag:true")
        if item.get("spoilerTrigger"):
            parts.append(f"spoilerTrigger:{js_quote(item['spoilerTrigger'])}")
        text = "  { " + ", ".join(parts) + " }"
        if i < len(locations) - 1:
            text += ","
        lines.append(text)
    lines.extend([
        "];",
        "",
        "var MAP_LOCATION_TYPES = " + js_quote(sorted({loc["type"] for loc in locations})) + ";",
        "",
    ])
    return "\n".join(lines)


def render_map_connections(connections: list[dict[str, Any]]) -> str:
    def js_number(v: Any) -> str:
        try:
            num = float(v)
        except Exception:
            return "0"
        if num.is_integer():
            return str(int(num))
        return format(num, "g")

    lines = [
        "// data/map_connections.js",
        "// Map connection lines synced from The Last Caretaker - Map Locations.xlsx (CONNECTIONS sheet).",
        "",
        "var MAP_CONNECTIONS = [",
    ]
    for i, item in enumerate(connections):
        text = (
            "  { id:" + str(int(item.get("id", 0) or 0))
            + ", from:" + js_quote(item.get("from") or "")
            + ", fromX:" + js_number(item.get("fromX"))
            + ", fromY:" + js_number(item.get("fromY"))
            + ", to:" + js_quote(item.get("to") or "")
            + ", toX:" + js_number(item.get("toX"))
            + ", toY:" + js_number(item.get("toY"))
            + " }"
        )
        if i < len(connections) - 1:
            text += ","
        lines.append(text)
    lines.extend([
        "];",
        "",
    ])
    return "\n".join(lines)


def diff_check(path: Path, expected: str) -> bool:
    current = path.read_text(encoding="utf-8")
    return current == expected


def main() -> int:
    parser = argparse.ArgumentParser(description="Sync data tables from workbook")
    mode = parser.add_mutually_exclusive_group(required=True)
    mode.add_argument("--write", action="store_true", help="Write generated outputs to data/*.js")
    mode.add_argument("--check", action="store_true", help="Fail if data/*.js are out of sync")
    args = parser.parse_args()

    wb = openpyxl.load_workbook(WORKBOOK_PATH, data_only=True)
    foods = parse_foods(wb)
    memories = parse_memories(wb)
    professions = parse_professions(wb)
    map_locations = parse_map_locations(MAP_WORKBOOK_PATH)
    map_connections = parse_map_connections(MAP_WORKBOOK_PATH)
    item_ids = load_item_ids_registry()

    # IDs are persisted in resource/item_ids.json so links remain stable over time.
    assign_stable_ids(foods, "foods", item_ids)
    assign_stable_ids(memories, "memories", item_ids)
    assign_stable_ids(professions, "professions", item_ids)

    rendered = {
        FOODS_PATH: render_foods(foods),
        MEMORIES_PATH: render_memories(memories),
        PROFESSIONS_PATH: render_professions(professions),
        MAP_LOCATIONS_PATH: render_map_locations(map_locations),
        MAP_CONNECTIONS_PATH: render_map_connections(map_connections),
        ITEM_IDS_PATH: render_item_ids_registry(item_ids),
    }

    if args.write:
        for path, text in rendered.items():
            path.write_text(text, encoding="utf-8")
        print(f"Updated {len(rendered)} files from workbook.")
        return 0

    out_of_sync = [path for path, text in rendered.items() if not diff_check(path, text)]
    if out_of_sync:
        print("Out of sync files:")
        for path in out_of_sync:
            print(" -", path.relative_to(ROOT))
        return 1

    print("Data files are in sync with workbook.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
